"""Unit tests for tabular_writer.py — pure functions, no DB/network."""

import io
import json

from openpyxl import load_workbook

from backend.services.data_transfer import tabular_writer as writer


def _bundles():
    return [
        {
            "entity_type": "driver",
            "entity_id": "d1",
            "user": {"id": "d1", "full_name": "Jane Doe"},
            "driver_profile": {"id": "driver-1", "license_plate": "ABC123"},
            "rides": [{"id": "r1"}, {"id": "r2"}],
            "documents": [{"id": "doc1"}],
            "driver_insurance_periods": [{"period": 1}],
        },
        {
            "entity_type": "rider",
            "entity_id": "r1",
            "user": {"id": "r1", "full_name": "John Smith"},
            "driver_profile": {},
            "rides": [],
            "documents": [],
            "driver_insurance_periods": [],
        },
    ]


def test_write_csv_one_row_per_entity_with_counts():
    csv_bytes = writer.write_csv(_bundles())
    text = csv_bytes.decode("utf-8-sig")
    assert "entity_type" in text
    assert "d1" in text and "r1" in text
    assert "rides_count" in text
    lines = [line for line in text.splitlines() if line.strip()]
    assert len(lines) == 3  # header + 2 entities


def test_write_csv_empty_bundle_list_returns_empty_bytes():
    assert writer.write_csv([]) == b""


def test_write_csv_sanitizes_formula_injection():
    """Same OWASP CSV-injection guard as the frontend's sanitizeCsvCell —
    a value starting with '=' must be neutralized with a leading quote."""
    bundles = [
        {
            "entity_type": "driver",
            "entity_id": "d1",
            "user": {"id": "d1", "full_name": "=cmd|'/c calc'!A1"},
            "driver_profile": {},
            "rides": [],
            "documents": [],
            "driver_insurance_periods": [],
        }
    ]
    csv_bytes = writer.write_csv(bundles)
    text = csv_bytes.decode("utf-8-sig")
    assert "'=cmd" in text  # prefixed with a literal apostrophe, not a bare formula


def test_write_json_round_trips_full_bundle_structure():
    json_bytes = writer.write_json(_bundles())
    parsed = json.loads(json_bytes)
    assert len(parsed) == 2
    assert parsed[0]["entity_id"] == "d1"
    assert parsed[0]["rides"] == [{"id": "r1"}, {"id": "r2"}]


def test_write_excel_produces_a_readable_workbook_with_header_and_rows():
    xlsx_bytes = writer.write_excel(_bundles())
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["entities"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 entities
    header = rows[0]
    assert "entity_id" in header
    entity_id_idx = header.index("entity_id")
    ids_in_rows = {rows[1][entity_id_idx], rows[2][entity_id_idx]}
    assert ids_in_rows == {"d1", "r1"}


def test_write_excel_empty_bundle_list_produces_header_only_sheet():
    xlsx_bytes = writer.write_excel([])
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["entities"]
    assert list(ws.iter_rows(values_only=True)) == []
